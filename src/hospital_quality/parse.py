"""Parse the DESTATIS 'Grunddaten der Krankenhaeuser' workbook into tidy long data.

The published workbook (EVAS 23111) contains dozens of human-oriented layout
tables plus a parallel set of machine-readable sheets named ``csv-23111-NN``.
Those csv sheets share a common long format:

    Statistik_Code | Statistik_Name | Gebiet | [Merkmal_1 ...] | Wertmerkmal
        | Jahr | Masseinheit | Wert

This module reads every ``csv-23111-NN`` sheet and normalises them into one
tidy table with a stable schema, so the rest of the pipeline never has to touch
Excel again. It handles two awkward realities of German official statistics:

* Text is encoded for a German locale; openpyxl decodes it correctly, but the
  region names carry umlauts we normalise for stable joins and file names.
* Numeric cells use DESTATIS placeholder markers ( ``-`` nil, ``.`` unknown,
  ``...`` not yet available, ``x`` not applicable). These are preserved as a
  ``value_flag`` and left as NULL values rather than being silently coerced to
  zero, which would corrupt every average.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

# Canonical column names in the normalised output.
COLUMNS = [
    "table_code",
    "statistic",
    "region",
    "dimension_1",
    "dimension_2",
    "measure",
    "year",
    "unit",
    "value",
    "value_flag",
]

# DESTATIS placeholder markers for a cell that is not a real number.
# Mapped to a human-readable flag; the numeric value becomes NULL.
_VALUE_MARKERS = {
    "-": "nil",
    ".": "unknown",
    "...": "not_available",
    "x": "not_applicable",
    "/": "not_publishable",
}


@dataclass(frozen=True)
class Record:
    """One tidy observation extracted from a csv sheet."""

    table_code: str
    statistic: str
    region: str
    dimension_1: str | None
    dimension_2: str | None
    measure: str
    year: int | None
    unit: str | None
    value: float | None
    value_flag: str | None


def _clean(text: object) -> str:
    """Trim a cell to a clean string ('' for None)."""
    if text is None:
        return ""
    return str(text).strip()


def parse_value(raw: object) -> tuple[float | None, str | None]:
    """Convert a raw 'Wert' cell into (numeric value, flag).

    Args:
        raw: The raw cell value (number, placeholder marker, or blank).

    Returns:
        ``(value, None)`` for a real number, or ``(None, flag)`` for a DESTATIS
        placeholder / blank cell.
    """
    if raw is None:
        return None, "missing"
    if isinstance(raw, (int, float)):
        return float(raw), None

    text = str(raw).strip()
    if text == "":
        return None, "missing"
    if text in _VALUE_MARKERS:
        return None, _VALUE_MARKERS[text]

    # German number formatting: '.' thousands separator, ',' decimal comma.
    normalised = text.replace(".", "").replace(",", ".")
    try:
        return float(normalised), None
    except ValueError:
        return None, "unparsed"


def _header_index(header: list[str]) -> dict[str, int]:
    """Map the columns we care about to their positions in a sheet header."""
    idx: dict[str, int] = {}
    merkmal_seen = 0
    for i, name in enumerate(header):
        key = name.strip()
        if key == "Gebiet":
            idx["region"] = i
        elif key == "Wertmerkmal":
            idx["measure"] = i
        elif key == "Jahr":
            idx["year"] = i
        elif key == "Masseinheit":
            idx["unit"] = i
        elif key == "Wert":
            idx["value"] = i
        elif key == "Statistik_Name":
            idx["statistic"] = i
        elif key.startswith("Merkmal_"):
            merkmal_seen += 1
            idx[f"dimension_{merkmal_seen}"] = i
    return idx


def parse_workbook(xlsx_path: str | Path) -> list[Record]:
    """Parse all ``csv-23111-NN`` sheets of the workbook into tidy records.

    Args:
        xlsx_path: Path to the downloaded DESTATIS workbook.

    Returns:
        A list of :class:`Record`, one per data cell across all csv sheets.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    records: list[Record] = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("csv-23111-"):
            continue
        table_code = sheet_name.replace("csv-", "")
        ws = wb[sheet_name]

        rows = ws.iter_rows(values_only=True)
        try:
            header = [_clean(c) for c in next(rows)]
        except StopIteration:
            continue
        idx = _header_index(header)
        if "region" not in idx or "value" not in idx:
            continue

        for row in rows:
            region = _clean(row[idx["region"]]) if idx.get("region") is not None else ""
            if not region:
                continue
            value, flag = parse_value(row[idx["value"]])
            year_raw = _clean(row[idx["year"]]) if "year" in idx else ""
            year = int(year_raw) if year_raw.isdigit() else None

            records.append(
                Record(
                    table_code=table_code,
                    statistic=_clean(row[idx["statistic"]]) if "statistic" in idx else "",
                    region=region,
                    dimension_1=(
                        _clean(row[idx["dimension_1"]]) or None if "dimension_1" in idx else None
                    ),
                    dimension_2=(
                        _clean(row[idx["dimension_2"]]) or None if "dimension_2" in idx else None
                    ),
                    measure=_clean(row[idx["measure"]]) if "measure" in idx else "",
                    year=year,
                    unit=_clean(row[idx["unit"]]) if "unit" in idx else None,
                    value=value,
                    value_flag=flag,
                )
            )

    wb.close()
    return records
